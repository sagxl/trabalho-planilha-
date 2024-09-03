from openpyxl import Workbook, load_workbook
import os     

# Verificação se o arquivo existe.
if not os.path.exists('configuracao_computadoresr.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.title = "Configuração de Computadores"

    #  cabeçalho da planilha 
    ws['A1'] = "ID"
    ws['B1'] = "Local"
    ws['C1'] = "Processador"
    ws['D1'] = "Memória RAM"
    ws['E1'] = "Disco Rígido"
    ws['F1'] = "Sistema Operacional" 

    # Dados dos computadores - tupla 
    escritorio_dados = [
        ("PC1", "Escritorio", "Intel i7", "8GB", "256GB SSD", "Windows 11"),
        ("PC2", "Escritorio", "Intel i7", "16GB", "512GB SSD", "Windows 11"),
        ("PC3", "Escritorio", "AMD Ryzen 7", "8GB", "256GB SSD", "Windows 11"),
        ("PC4", "Escritorio", "Intel i7", "8GB", "256GB SSD", "Linux"),
        ("PC5", "Escritorio", "AMD Ryzen 7", "16GB", "256GB SSD", "Linux"),
        ("PC6", "Escritorio", "Intel i7", "16GB", "512GB SSD", "Windows 11"),
        ("PC7", "Escritorio", "AMD Ryzen 7", "16GB", "512GB SSD", "Linux"),
        ("PC8", "Escritorio", "AMD Ryzen 7", "8GB", "256GB SSD", "Linux"),
        ("PC9", "Escritorio", "Intel i7", "8GB", "512GB SSD", "Linux"),
        ("PC10", "Escritorio", "AMD Ryzen 7", "16GB", "512GB SSD", "Windows 11")
    ]

    # Adiciona os dados - tupla 
    for i, data in enumerate(escritorio_dados, start=2): # 
        for j, value in enumerate(data, start=1):
            ws.cell(row=i, column=j, value=value)

    wb.save('configuracao_computadoresr.xlsx')

# Carrega o workbook 
book = load_workbook('configuracao_computadoresr.xlsx')
sheet = book.active


ultimo_id_cadastrado = None


print('Opção 1 - Solicitar ao usuário uma busca nas linhas dos PCs (1 a 10)')
print('Opção 2 - Identificar quais dos PCs atendem aos critérios de RAM, Sistema Operacional, Processador e Disco Rígido')
print('Opção 3 - Fazer um cadastro de um novo PC')
opcao = input('Escolha uma opção [1] [2] [3]: ')

if opcao == '1':
   
    if ultimo_id_cadastrado:
        linha_id = ultimo_id_cadastrado
    else:
        linha_id = input('Escolha um ID de linha (por exemplo, PC1): ')

    resultado = None
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=6):
        if row[0].value == linha_id:
            resultado = row
            break

    if resultado:
        print("------------------------------------------------")
        print(f"{resultado[0].value} {resultado[1].value} {resultado[2].value} {resultado[3].value} {resultado[4].value} {resultado[5].value}")
    else:
        print("Linha não encontrada.")

elif opcao == '2':
    
    book = load_workbook('configuracao_computadoresr.xlsx')
    sheet = book.active

    criterios = {
        "Windows 10": "Windows 10",
        "Windows 11": "Windows 11",
        "Linux": "Linux",
        "4GB RAM": "4GB",
        "8GB RAM": "8GB",
        "16GB RAM": "16GB",
        "32GB RAM": "32GB",
        "64GB RAM": "64GB",
        "Intel i5": "Intel i5",
        "Intel i7": "Intel i7",
        "Intel i9": "Intel i9",
        "AMD Ryzen 5": "AMD Ryzen 5",
        "AMD Ryzen 7": "AMD Ryzen 7",
        "AMD Ryzen 9": "AMD Ryzen 9",
        "256GB HDD": "256GB HDD",
        "512GB HDD": "512GB HDD",
        "1T HDD": "1T HDD",
        "2T HDD": "2T HDD",
        "256GB SSD": "256GB SSD",
        "512GB SSD": "512GB SSD",
        "1T SSD": "1T SSD",
        "2T SSD": "2T SSD"
    }

    for criterio, valor in criterios.items():
        print(f"PCs que atendem ao critério '{criterio}':")
        print("------------------------------------------------")
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=6):
            if (criterio in ["Windows 11", "Windows 10", "Linux"] and row[5].value == valor) or \
               (criterio in ["4GB RAM", "8GB RAM", "16GB RAM", "32GB RAM", "64GB RAM"] and row[3].value == valor) or \
               (criterio in ["Intel i5", "Intel i7", "Intel i9", "AMD Ryzen 5", "AMD Ryzen 7", "AMD Ryzen 9"] and row[2].value == valor) or \
               (criterio in ["256GB HDD", "256GB SSD", "512GB HDD", "512GB SSD", "1T HDD", "1T SSD", "2T HDD", "2T SSD"] and row[4].value == valor):
                print(f"{row[0].value} {row[1].value} {row[2].value} {row[3].value} {row[4].value} {row[5].value}")
        print("------------------------------------------------")

elif opcao == '3':
    print("Cadastro de novo PC:")
    id_pc = input("ID do PC: ")
    local = input("Local: ")
    processador = input("Processador (Intel i7 ou AMD Ryzen 7): ")
    memoria = input("Memória RAM (8GB ou 16GB): ")
    disco = input("Disco Rígido (256GB SSD ou 512GB SSD): ")
    sistema = input("Sistema Operacional (Windows 11 ou Linux): ")

    # Adicionar os dados 
    sheet.append([id_pc, local, processador, memoria, disco, sistema])
    book.save('configuracao_computadoresr.xlsx')
    print("Novo PC cadastrado !")

    
    ultimo_id_cadastrado = id_pc

    # vai Recarrega o workbook e a planilha para mostrar os novos dados 
    book = load_workbook('configuracao_computadoresr.xlsx')
    sheet = book.active

else:
    print("Opção inválida. Escolha [1], [2] ou [3].")